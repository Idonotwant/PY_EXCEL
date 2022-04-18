from openpyxl import Workbook
from os import sep

wb = Workbook()
ws = wb.active

targetPos = input("Please input pos(absolute): ")
name = input("please input name: ")
targetPos = targetPos[1:-1]+sep+name+'.xlsx'
ws['A1'].value = 'Title'
ws['A2'].value = 'xStart'
ws['A3'].value = 'xend'
ws['A4'].value = 'ystart'
ws['A5'].value = 'yend'
ws['A6'].value = 'chartPos'

wb.save(targetPos)