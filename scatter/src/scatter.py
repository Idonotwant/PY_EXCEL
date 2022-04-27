'''
Author: Wei-Ju-Chen(Idonotwant)
ProjectName: scatter
Description: scatter
Brief_Log:
2022/04/16 Start
2022/04/18 v1 packed
2022/04/27 class form
Note: 
python setting.py to make reference file
python scatter.py to make scatter with log x axis(is usually used in frequence response)
'''
from openpyxl import load_workbook,drawing
from os import system,sep
from sys import exit
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
    marker
)

blue = drawing.colors.ColorChoice(prstClr='blue')

class scatter():
    title = 'B1'
    xlabel = 'B2'
    xend = 'B3'
    ylabel = 'B4'
    yend = 'B5'
    chartPos = 'B6'
    def __init__(self,**kwgs):
        if 'name' in  kwgs.keys():
            self.name = kwgs['name']
        else:
            self.name = 'newScatter'
        if 'mode' in kwgs.keys():
            self.mode = kwgs['mode']
        else:
            self.mode = 0
        self.bookpass = self.get(self.mode)
        self.read_plot()
    def get(self,mode):
        if not mode:
            #input by typing
            bookpass = input('worbook poisition(absolute: ')
            #debug
            bookpass = bookpass[1:-1]
            return bookpass
        else:#input by tkinter
            pass
    def read_plot(self):
        wb = load_workbook(self.bookpass)
        ws = wb.active
        chart = ScatterChart(scatterStyle='marker')
        #scatterStyle=None,varyColors=NONE,dLbls=NONE,extLst=None
        chart.title = ws[ws[scatter.title].value].value
        chart.x_axis.title = ws[ws[scatter.xlabel].value].value
        chart.y_axis.title = ws[ws[scatter.ylabel].value].value
        chart.x_axis.scaling.logBase = 10

        x_min_col,x_min_row = self.splitter(ws[scatter.xlabel].value)
        x_min_row+=1
        x_max_col,x_max_row = self.splitter(ws[scatter.xend].value)

        y_min_col,y_min_row = self.splitter(ws[scatter.ylabel].value)
        y_min_row+=1
        y_max_col,y_max_row = self.splitter(ws[scatter.yend].value)

        xvalues = Reference(ws, min_col=x_min_col, min_row=x_min_row, max_row=x_max_row)
        values = Reference(ws, min_col=y_min_col, min_row=y_min_row, max_row=y_max_row)
        series = Series(values, xvalues, title_from_data=False)

        chartPosV = ws[scatter.chartPos].value

        lineProp = drawing.line.LineProperties(solidFill = blue)
        series.graphicalProperties.line = lineProp
        chart.series.append(series)

        chart.style = 25
        ws.add_chart(chart,chartPosV)

        tmp = self.bookpass.split('.')
        bookoutpass = tmp[0]+"_out."+tmp[1]
        wb.save(bookoutpass)
    def find_col(self,col): #ex: "A"->1
        col = col.upper()
        col = "".join(reversed(col))
        base = 1
        colValue = 0
        for i in range (len(col)):
            colValue += base*(ord(col[i])-64)
            base *= 26
        return colValue
    def splitter(self,target): #ex: "BA123"->53,123
        i = 0
        while target[i].isalpha():
            i+=1
        return self.find_col(target[:i]) ,int(target[i:])



if __name__ == "__main__":
    a = scatter()

