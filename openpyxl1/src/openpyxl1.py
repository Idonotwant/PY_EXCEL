'''
Author: Wei-Ju-Chen(Idonotwant)
ProjectName: openpyxl1
Description: openpyxl1
Brief_Log:
2022/04/16 Start
Note: 
intro to openpyxl
'''

from datetime import datetime
from openpyxl import Workbook

wb = Workbook()

ws = wb.active

ws['A1'] = 42

ws.append([1,2,3])


ws['A2'] = datetime.now()

wb.save("test1.xlsx")